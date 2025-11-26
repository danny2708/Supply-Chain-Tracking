import io
from datetime import datetime, date  # SỬA: Import thêm date
from typing import List, Dict, Tuple

import requests
from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers, Alignment # Import numbers để format date trong excel

from rest_framework import status
from rest_framework.response import Response

from dateutil import parser as date_parser 

def _try_parse_date(v: str):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date): # Handle trường hợp đã là date
        return v
    v = str(v).strip()
    # try ISO or flexible parsing
    try:
        d = date_parser.parse(v, dayfirst=True)
        return d.date()
    except Exception:
        # fallback - try common formats
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except Exception:
                continue
    return None


def parse_excel_and_create_products(uploaded_file, request, ProductSerializer):
    """
    Parse uploaded Excel file (xlsx) and create products.
    """
    created = []
    errors = []

    data = uploaded_file.read()
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = [str(c).strip() if c is not None else "" for c in row]
        break

    if not header_row:
        return created, [{"row": 0, "errors": "Empty header"}]

    header_map = {h.lower(): idx for idx, h in enumerate(header_row)}

    def gv(row_values, key):
        idx = header_map.get(key)
        if idx is None:
            return None
        v = row_values[idx] if idx < len(row_values) else None
        return v

    row_number = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_number += 1
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        product_id = gv(row, "product_id") or gv(row, "id") or ""
        name = gv(row, "name") or ""
        description = gv(row, "description") or ""
        manufacture_date_raw = gv(row, "manufacture_date") or gv(row, "manufactured") or ""
        expiry_date_raw = gv(row, "expiry_date") or gv(row, "expiry") or ""
        ipfs = gv(row, "ipfs") or ""

        manufacture_date = _try_parse_date(manufacture_date_raw)
        expiry_date = _try_parse_date(expiry_date_raw)

        payload = {
            "product_id": str(product_id).strip(),
            "name": str(name).strip(),
            "description": description if description is not None else "",
            "manufacture_date": manufacture_date.isoformat() if manufacture_date else None,
            "expiry_date": expiry_date.isoformat() if expiry_date else None,
            "ipfs": str(ipfs).strip() if ipfs else "",
        }

        serializer = ProductSerializer(data=payload, context={"request": request})
        try:
            serializer.is_valid(raise_exception=True)
            product = serializer.save()
            created.append(ProductSerializer(product, context={"request": request}).data)
        except Exception as e:
            err_detail = getattr(e, "detail", None)
            if err_detail is None:
                err_detail = str(e)
            errors.append({"row": row_number, "errors": err_detail, "row_data": payload})

    return created, errors


def build_excel_from_products_list(products_list: List[Dict]) -> bytes:
    """
    Build Excel bytes from list of dicts with FORMATTING:
    - Left align default
    - Wrap text
    - Fixed width for IPFS column
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "product_id", "name", "description", 
        "manufacture_date", "expiry_date", "ipfs", 
        "user", "username", "on_chain_status",
    ]
    ws.append(headers)

    # Helper parse
    def parse_date(d):
        if not d: return ""
        try:
            return datetime.strptime(str(d)[:10], "%Y-%m-%d").date() # Safe slice for ISO
        except:
            return d 

    for p in products_list:
        m_date = parse_date(p.get("manufacture_date"))
        e_date = parse_date(p.get("expiry_date"))

        row = [
            p.get("product_id", ""),
            p.get("name", ""),
            p.get("description", ""),
            m_date,
            e_date,
            p.get("ipfs", ""),
            p.get("user", ""),
            p.get("username", ""),
            p.get("on_chain_status", ""),
        ]
        ws.append(row)

    # ------------ FORMATTING STYLES ------------
    
    # 1. Tạo style: Căn trái, lên trên, tự động xuống dòng (wrap text)
    align_style = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 2. Áp dụng format cho TOÀN BỘ các ô (bao gồm cả header nếu muốn)
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = align_style

    # 3. Format Date (DD/MM/YYYY)
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5): 
        for cell in row:
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "DD/MM/YYYY"

    # 4. Column Widths logic (Xử lý độ rộng cột)
    for column_cells in ws.columns:
        # Lấy header của cột hiện tại (dòng 1)
        header_cell = column_cells[0] 
        col_letter = header_cell.column_letter
        header_name = str(header_cell.value).lower() if header_cell.value else ""

        # Xử lý riêng cho cột IPFS: Giới hạn độ rộng cố định
        if "ipfs" in header_name:
            ws.column_dimensions[col_letter].width = 35 # Đủ rộng để xem, đủ hẹp để wrap text
        
        elif "description" in header_name:
             ws.column_dimensions[col_letter].width = 40 
             
        else:
            # Logic Auto-fit cho các cột còn lại
            max_length = 0
            for cell in column_cells:
                try:
                    val = cell.value
                    if val:
                        # Giới hạn max_length = 50 để không bị cột quá to
                        line_len = len(str(val))
                        if line_len > 50: line_len = 50 
                        max_length = max(max_length, line_len)
                except:
                    pass
            # Set width + một chút padding
            ws.column_dimensions[col_letter].width = max_length + 2

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def fetch_products_from_api(api_url: str, auth_header: str, params: dict = None):
    headers = {}
    if auth_header:
        headers["Authorization"] = auth_header
    # Thêm verify=False nếu chạy localhost https self-signed, nếu http thường thì ko cần
    resp = requests.get(api_url, headers=headers, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()